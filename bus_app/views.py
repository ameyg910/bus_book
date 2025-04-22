from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Bus, Booking, User, SeatClass, Stop, Passenger
from django.utils import timezone
from datetime import timedelta
from django.contrib.auth import login as auth_login
from django.contrib.auth.views import LoginView
from django.shortcuts import render, redirect
from django.contrib.auth import login as auth_login
from .forms import SignUpForm
from django.contrib.auth.views import LogoutView
from decimal import Decimal
from django.db.models import Q
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
import logging
from .utils import send_verification_email, send_otp_email
from .utils import send_booking_confirmation_email
from .models import OTP 
from django.urls import reverse

def is_admin(user):
    return user.is_admin

@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    buses = Bus.objects.all()
    bookings = Booking.objects.all()
    return render(request, 'bus_app/admin_dashboard.html', {'buses': buses, 'bookings': bookings})

@login_required
@user_passes_test(is_admin)
def add_bus(request):
    if request.method == 'POST':
        bus_name = request.POST['bus_name']
        days_of_week = request.POST['days_of_week']
        bus = Bus(bus_name=bus_name, days_of_week=days_of_week)
        bus.save()

        # Add seat classes
        for class_type in ['General', 'Sleeper', 'Luxury']:
            total_seats = int(request.POST.get(f'{class_type.lower()}_seats', 0))
            if total_seats > 0:
                SeatClass.objects.create(
                    bus=bus,
                    class_type=class_type,
                    total_seats=total_seats,
                    available_seats=total_seats,
                    fare=request.POST[f'{class_type.lower()}_fare']
                )

        # Add stops
        stop_locations = request.POST.getlist('stop_location')
        stop_times = request.POST.getlist('stop_time')
        for i, (location, time) in enumerate(zip(stop_locations, stop_times), 1):
            Stop.objects.create(bus=bus, location=location, arrival_time=time, sequence=i)

        return redirect('admin_dashboard')
    return render(request, 'bus_app/add_bus.html')

@login_required
@user_passes_test(is_admin)
def cancel_bus(request, bus_id):
    bus = get_object_or_404(Bus, id=bus_id)
    bookings = Booking.objects.filter(bus=bus)
    for booking in bookings:
        booking.user.wallet_balance += booking.total_cost
        booking.user.save()
        booking.delete()
    bus.delete()
    return redirect('admin_dashboard')

@login_required
def search_buses(request):
    if request.method == 'POST':
        source = request.POST['source']
        destination = request.POST['destination']
        buses = Bus.objects.filter(
            stops__location=source
        ).filter(
            stops__location=destination
        ).distinct()
        valid_buses = []
        for bus in buses:
            source_stop = bus.stops.get(location=source)
            dest_stop = bus.stops.get(location=destination)
            if source_stop.sequence < dest_stop.sequence:
                valid_buses.append({
                    'bus': bus,
                    'source_stop': source_stop,
                    'dest_stop': dest_stop,
                    'seat_classes': bus.seat_classes.all()
                })
        return render(request, 'bus_app/search_results.html', {'valid_buses': valid_buses})
    return render(request, 'bus_app/search_buses.html')

logger = logging.getLogger(__name__)

@login_required
def book_ticket(request, bus_id):
    bus = get_object_or_404(Bus, id=bus_id)
    if request.method == 'POST':
        seat_class_id = request.POST.get('seat_class')
        start_stop_id = request.POST.get('start_stop')
        end_stop_id = request.POST.get('end_stop')
        passenger_names = request.POST.getlist('passenger_name')
        passenger_ages = request.POST.getlist('passenger_age')

        try:
            seat_class = SeatClass.objects.get(id=seat_class_id)
            start_stop = Stop.objects.get(id=start_stop_id)
            end_stop = Stop.objects.get(id=end_stop_id)
            total_cost = seat_class.fare * len(passenger_names)

            if seat_class.available_seats >= len(passenger_names) and request.user.wallet_balance >= total_cost:
                booking = Booking.objects.create(
                    user=request.user, bus=bus, seat_class=seat_class,
                    total_cost=total_cost, start_stop=start_stop, end_stop=end_stop
                )
                for name, age in zip(passenger_names, passenger_ages):
                    passenger = Passenger.objects.create(name=name, age=age)
                    booking.passengers.add(passenger)
                seat_class.available_seats -= len(passenger_names)
                seat_class.save()
                request.user.wallet_balance -= total_cost
                request.user.save()
                send_booking_confirmation_email(booking)  # Send confirmation email
                send_otp_email(request.user, 'booking')  # Send OTP
                return redirect('verify_otp', purpose='booking', user_id=request.user.id)
            else:
                return render(request, 'bus_app/book_ticket.html', {
                    'bus': bus,
                    'error': "Not enough seats or insufficient wallet balance."
                })
        except (SeatClass.DoesNotExist, Stop.DoesNotExist) as e:
            return render(request, 'bus_app/book_ticket.html', {'bus': bus, 'error': str(e)})
    return render(request, 'bus_app/book_ticket.html', {'bus': bus})

@login_required
def my_trips(request):
    bookings = Booking.objects.filter(user=request.user)
    return render(request, 'bus_app/my_trips.html', {'bookings': bookings})

@login_required
def cancel_trip(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id, user=request.user)
    if timezone.now() < booking.bus.departure_time - timedelta(hours=6):
        booking.bus.available_seats += booking.passengers.count()
        booking.bus.save()
        request.user.wallet_balance += booking.total_cost
        request.user.save()
        booking.delete()
    return redirect('my_trips')
@login_required
def add_funds(request):
    if request.method == 'POST':
        try:
            amount = Decimal(request.POST['amount'])  # Convert to Decimal instead of float
            if amount <= 0:
                return render(request, 'bus_app/add_funds.html', {'error': 'Amount must be positive'})
            request.user.wallet_balance += amount
            request.user.save()
            return redirect('my_trips')
        except ValueError:
            return render(request, 'bus_app/add_funds.html', {'error': 'Please enter a valid number'})
        except Exception as e:
            return render(request, 'bus_app/add_funds.html', {'error': f'An error occurred: {str(e)}'})
    return render(request, 'bus_app/add_funds.html')

class CustomLoginView(LoginView):
    template_name = 'registration/login.html'
    authentication_form = None

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            if request.user.is_staff:  # Changed from is_admin to is_staff
                return redirect('/admin/')
            return redirect('search_buses')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        auth_login(self.request, form.get_user())
        if form.get_user().is_staff:  # Changed from is_admin to is_staff
            return redirect('/admin/')
        return redirect('search_buses')

    def get_success_url(self):
        if self.request.user.is_staff:  # Changed from is_admin to is_staff
            return '/admin/'
        return reverse('search_buses')

def signup(request):
    if request.user.is_authenticated:
        print("User already authenticated, redirecting to home")
        return redirect('/')
    if request.method == 'POST':
        print("Signup POST received")
        form = SignUpForm(request.POST)
        if form.is_valid():
            print("Form is valid")
            user = form.save(commit=False)
            user.is_passenger = True
            user.save()
            print("User saved:", user.username)
            send_verification_email(user)
            print("Verification email sent")
            send_otp_email(user, 'signup')
            print("OTP email sent")
            return redirect('verify_otp', purpose='signup', user_id=user.id)
    else:
        form = SignUpForm()
    return render(request, 'bus_app/signup.html', {'form': form})

class CustomLogoutView(LogoutView):
    template_name = 'registration/logged_out.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        return super().dispatch(request, *args, **kwargs)

@login_required
@user_passes_test(is_admin)
def bus_bookings(request, bus_id):
    bus = get_object_or_404(Bus, id=bus_id)
    bookings = Booking.objects.filter(bus=bus)
    return render(request, 'bus_app/bus_bookings.html', {'bus': bus, 'bookings': bookings})

@login_required
@user_passes_test(is_admin)
def update_bus(request, bus_id):
    bus = get_object_or_404(Bus, id=bus_id)
    if request.method == 'POST':
        bus.bus_name = request.POST['bus_name']
        bus.source = request.POST['source']
        bus.destination = request.POST['destination']
        bus.fare = request.POST['fare']
        bus.total_seats = request.POST['total_seats']
        bus.available_seats = min(bus.available_seats, int(request.POST['total_seats']))
        bus.departure_time = request.POST['departure_time']
        bus.days_of_week = request.POST['days_of_week']
        bus.save()
        return redirect('admin_dashboard')
    return render(request, 'bus_app/update_bus.html', {'bus': bus})

@login_required
def edit_booking_passengers(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id, user=request.user)
    if request.method == 'POST':
        for passenger in booking.passengers.all():
            passenger.name = request.POST.get(f'passenger_name_{passenger.id}')
            passenger.age = request.POST.get(f'passenger_age_{passenger.id}')
            passenger.save()
        return redirect('my_trips')
    return render(request, 'bus_app/edit_booking_passengers.html', {'booking': booking})

@login_required
@user_passes_test(is_admin)
def export_bookings(request, bus_id):
    bus = get_object_or_404(Bus, id=bus_id)
    bookings = Booking.objects.filter(bus=bus)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Bookings for {bus.bus_name}"

    headers = ['User', 'Seat Class', 'Total Cost', 'Start Stop', 'End Stop', 'Passengers']
    for col_num, header in enumerate(headers, 1):
        ws[f'{get_column_letter(col_num)}1'] = header

    for row_num, booking in enumerate(bookings, 2):
        ws[f'A{row_num}'] = str(booking.user)
        ws[f'B{row_num}'] = booking.seat_class.class_type
        ws[f'C{row_num}'] = float(booking.total_cost)
        ws[f'D{row_num}'] = booking.start_stop.location
        ws[f'E{row_num}'] = booking.end_stop.location
        ws[f'F{row_num}'] = ', '.join([p.name for p in booking.passengers.all()])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="bookings_{bus.bus_name}.xlsx"'
    wb.save(response)
    return response

def verify_otp(request, purpose, user_id):
    user = User.objects.get(id=user_id)
    if request.method == 'POST':
        otp_input = request.POST.get('otp')
        try:
            otp_record = OTP.objects.filter(user=user, purpose=purpose).latest('created_at')
            if not otp_record.is_expired() and otp_record.otp == otp_input:
                if purpose == 'signup':
                    user.email_verified = True  # Mark email as verified after OTP
                    user.save()
                    return redirect('login')
                elif purpose == 'booking':
                    return render(request, 'bus_app/my_trips.html', {
                        'bookings': Booking.objects.filter(user=user),
                        'success': "Booking confirmed!"
                    })
            else:
                return render(request, 'bus_app/verify_otp.html', {
                    'error': 'Invalid or expired OTP',
                    'purpose': purpose,
                    'user_id': user_id
                })
        except OTP.DoesNotExist:
            return render(request, 'bus_app/verify_otp.html', {
                'error': 'No OTP found',
                'purpose': purpose,
                'user_id': user_id
            })
    return render(request, 'bus_app/verify_otp.html', {'purpose': purpose, 'user_id': user_id})

def verify_email(request, token):
    try:
        user = User.objects.get(email_verification_token=token)
        if timezone.now() < user.email_verification_expiry:
            user.email_verified = True
            user.email_verification_token = None
            user.email_verification_expiry = None
            user.save()
            return render(request, 'bus_app/message.html', {'message': 'Email verified successfully! Please log in.'})
        else:
            return render(request, 'bus_app/message.html', {'message': 'Verification link has expired.'})
    except User.DoesNotExist:
        return render(request, 'bus_app/message.html', {'message': 'Invalid verification link.'})

@login_required
def wallet(request):
    if request.method == 'POST':
        amount = Decimal(request.POST.get('amount', 0))
        if amount > 0:
            request.user.wallet_balance += amount
            request.user.save()
            return redirect('wallet')
    return render(request, 'bus_app/wallet.html')

@login_required
def profile(request):
    if request.method == 'POST':
        user = request.user
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        user.phone_number = request.POST.get('phone_number', user.phone_number)
        user.save()
        return redirect('profile')
    return render(request, 'bus_app/profile.html')